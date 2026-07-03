import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import io
import sys
import shutil
from PyPDF2 import PdfMerger
import threading
from PIL import Image, ImageTk
import tempfile

try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    import pytesseract
    PYTESSERACT_AVAILABLE = True
except ImportError:
    PYTESSERACT_AVAILABLE = False

try:
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    PDF_EXCEL_AVAILABLE = True
except ImportError:
    PDF_EXCEL_AVAILABLE = False


class PDFMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Ferramentas PDF")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # Variáveis - Juntar PDFs
        self.pasta_origem = tk.StringVar()
        self.arquivo_destino = tk.StringVar()
        self.arquivos_pdf = []
        self.preview_image = None
        self.preview_token = 0
        self.preview_cache = {}

        # Variáveis - OCR
        self.ocr_arquivo_entrada = tk.StringVar()
        self.ocr_arquivo_saida = tk.StringVar()

        # Variáveis - Comprimir PDF
        self.compress_arquivo_entrada = tk.StringVar()
        self.compress_arquivo_saida = tk.StringVar()
        self.compress_nivel = tk.StringVar(value="media")

        # Variáveis - PDF para Excel
        self.excel_arquivo_entrada = tk.StringVar()
        self.excel_arquivo_saida = tk.StringVar()
        self.excel_modo = tk.StringVar(value="tabelas")

        self.configurar_estilo()
        self.criar_interface()

        if not PDF2IMAGE_AVAILABLE:
            messagebox.showwarning(
                "Aviso",
                "A biblioteca pdf2image não está instalada.\n"
                "A pré-visualização de PDFs não estará disponível.\n"
                "Para instalar: pip install pdf2image"
            )

    def configurar_estilo(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Title.TLabel', font=('Arial', 18, 'bold'), foreground='#2c3e50')
        style.configure('Heading.TLabel', font=('Arial', 12, 'bold'), foreground='#34495e')
        style.configure('Custom.TButton', font=('Arial', 10))

    def criar_interface(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)

        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.criar_aba_juntar()
        self.criar_aba_ocr()
        self.criar_aba_comprimir()
        self.criar_aba_pdf_excel()

    # ── Aba: Juntar PDFs ────────────────────────────────────────────────────

    def criar_aba_juntar(self):
        aba = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(aba, text="🔗 Juntar PDFs")

        aba.columnconfigure(1, weight=1)
        aba.rowconfigure(3, weight=1)

        ttk.Label(aba, text="🔗 Juntar PDFs", style='Title.TLabel').grid(
            row=0, column=0, columnspan=3, pady=(0, 20))

        self.criar_secao_selecao(aba)

        files_frame = ttk.LabelFrame(aba, text="Gerenciamento de Arquivos", padding="15")
        files_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=20)
        files_frame.columnconfigure(0, weight=1)
        files_frame.columnconfigure(1, weight=1)
        files_frame.rowconfigure(1, weight=1)

        self.criar_secao_arquivos(files_frame)
        self.criar_secao_preview(files_frame)
        self.criar_secao_progresso(aba)

        self.btn_juntar = ttk.Button(aba, text="🔗 Juntar PDFs",
                                     command=self.juntar_pdfs_thread,
                                     state="disabled",
                                     style='Custom.TButton')
        self.btn_juntar.grid(row=6, column=0, columnspan=3, pady=20, ipadx=20, ipady=10)

    def criar_secao_selecao(self, parent):
        ttk.Label(parent, text="📁 Pasta com PDFs:", style='Heading.TLabel').grid(
            row=1, column=0, sticky=tk.W, pady=5)

        entry_pasta = ttk.Entry(parent, textvariable=self.pasta_origem,
                                state="readonly", width=50)
        entry_pasta.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)

        btn_pasta = ttk.Button(parent, text="Procurar",
                               command=self.selecionar_pasta)
        btn_pasta.grid(row=1, column=2, pady=5)

        ttk.Label(parent, text="💾 Salvar como:", style='Heading.TLabel').grid(
            row=2, column=0, sticky=tk.W, pady=5)

        entry_destino = ttk.Entry(parent, textvariable=self.arquivo_destino,
                                  state="readonly", width=50)
        entry_destino.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)

        btn_destino = ttk.Button(parent, text="Procurar",
                                 command=self.selecionar_destino)
        btn_destino.grid(row=2, column=2, pady=5)

    def criar_secao_arquivos(self, parent):
        left_frame = ttk.Frame(parent)
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        left_frame.columnconfigure(0, weight=1)
        left_frame.rowconfigure(1, weight=1)

        control_frame = ttk.Frame(left_frame)
        control_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        control_frame.columnconfigure(0, weight=1)

        ttk.Label(control_frame, text="📄 Arquivos PDF:", style='Heading.TLabel').grid(
            row=0, column=0, sticky=tk.W)

        btn_frame = ttk.Frame(control_frame)
        btn_frame.grid(row=0, column=1, sticky=tk.E)

        self.btn_adicionar = ttk.Button(btn_frame, text="➕ Adicionar",
                                        command=self.adicionar_arquivos)
        self.btn_adicionar.grid(row=0, column=0, padx=2)

        self.btn_remover = ttk.Button(btn_frame, text="➖ Remover",
                                      command=self.remover_selecionados,
                                      state="disabled")
        self.btn_remover.grid(row=0, column=1, padx=2)

        self.btn_subir = ttk.Button(btn_frame, text="⬆️",
                                    command=self.mover_para_cima,
                                    state="disabled", width=3)
        self.btn_subir.grid(row=0, column=2, padx=2)

        self.btn_descer = ttk.Button(btn_frame, text="⬇️",
                                     command=self.mover_para_baixo,
                                     state="disabled", width=3)
        self.btn_descer.grid(row=0, column=3, padx=2)

        list_frame = ttk.Frame(left_frame)
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        self.listbox = tk.Listbox(list_frame, height=12, selectmode=tk.EXTENDED)
        self.listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.listbox.bind('<<ListboxSelect>>', self.on_listbox_select)

        scrollbar_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.listbox.yview)
        scrollbar_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.listbox.configure(yscrollcommand=scrollbar_y.set)

        scrollbar_x = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.listbox.xview)
        scrollbar_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        self.listbox.configure(xscrollcommand=scrollbar_x.set)

    def criar_secao_preview(self, parent):
        right_frame = ttk.LabelFrame(parent, text="🔍 Pré-visualização", padding="10")
        right_frame.grid(row=0, column=1, rowspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(10, 0))
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(1, weight=1)

        self.preview_info = ttk.Label(right_frame, text="Selecione um arquivo para ver a pré-visualização")
        self.preview_info.grid(row=0, column=0, pady=(0, 10))

        self.preview_canvas = tk.Canvas(right_frame, bg='white', relief='sunken', bd=2)
        self.preview_canvas.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        preview_scroll_y = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.preview_canvas.yview)
        preview_scroll_y.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.preview_canvas.configure(yscrollcommand=preview_scroll_y.set)

        preview_scroll_x = ttk.Scrollbar(right_frame, orient=tk.HORIZONTAL, command=self.preview_canvas.xview)
        preview_scroll_x.grid(row=2, column=0, sticky=(tk.W, tk.E))
        self.preview_canvas.configure(xscrollcommand=preview_scroll_x.set)

    def criar_secao_progresso(self, parent):
        self.progress = ttk.Progressbar(parent, mode='determinate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)

        self.status_label = ttk.Label(parent, text="Selecione uma pasta com arquivos PDF ou adicione arquivos individuais")
        self.status_label.grid(row=5, column=0, columnspan=3, pady=5)

    # ── Aba: OCR ────────────────────────────────────────────────────────────

    def criar_aba_ocr(self):
        aba = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(aba, text="🔍 OCR - PDF Pesquisável")

        aba.columnconfigure(1, weight=1)

        ttk.Label(aba, text="🔍 OCR - PDF Pesquisável", style='Title.TLabel').grid(
            row=0, column=0, columnspan=3, pady=(0, 20))

        ttk.Label(aba, text="📄 PDF de entrada:", style='Heading.TLabel').grid(
            row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(aba, textvariable=self.ocr_arquivo_entrada, state="readonly", width=50).grid(
            row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(aba, text="Procurar", command=self.ocr_selecionar_entrada).grid(
            row=1, column=2, pady=5)

        ttk.Label(aba, text="💾 Salvar como:", style='Heading.TLabel').grid(
            row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(aba, textvariable=self.ocr_arquivo_saida, state="readonly", width=50).grid(
            row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(aba, text="Procurar", command=self.ocr_selecionar_saida).grid(
            row=2, column=2, pady=5)

        ttk.Label(aba,
                  text="ℹ️  Requer Tesseract OCR instalado  —  github.com/tesseract-ocr/tesseract",
                  foreground='#666666').grid(row=3, column=0, columnspan=3, pady=(15, 5))

        self.ocr_progress = ttk.Progressbar(aba, mode='determinate')
        self.ocr_progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)

        self.ocr_status_label = ttk.Label(aba, text="Selecione um PDF de imagem para converter em pesquisável")
        self.ocr_status_label.grid(row=5, column=0, columnspan=3, pady=5)

        self.btn_ocr = ttk.Button(aba, text="🔍 Gerar PDF com OCR",
                                   command=self.ocr_realizar_thread,
                                   state="disabled",
                                   style='Custom.TButton')
        self.btn_ocr.grid(row=6, column=0, columnspan=3, pady=30, ipadx=20, ipady=10)

    def ocr_selecionar_entrada(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o PDF de imagem",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )
        if arquivo:
            self.ocr_arquivo_entrada.set(arquivo)
            self._ocr_atualizar_botao()

    def ocr_selecionar_saida(self):
        arquivo = filedialog.asksaveasfilename(
            title="Salvar PDF pesquisável como",
            defaultextension=".pdf",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )
        if arquivo:
            self.ocr_arquivo_saida.set(arquivo)
            self._ocr_atualizar_botao()

    def _ocr_atualizar_botao(self):
        state = "normal" if (self.ocr_arquivo_entrada.get() and self.ocr_arquivo_saida.get()) else "disabled"
        self.btn_ocr.config(state=state)

    def ocr_realizar_thread(self):
        thread = threading.Thread(target=self.ocr_realizar)
        thread.daemon = True
        thread.start()

    def _ocr_status(self, texto=None, valor=None):
        self._status_ui(self.ocr_status_label, texto, self.ocr_progress, valor)

    def ocr_realizar(self):
        if not PDF2IMAGE_AVAILABLE:
            self.root.after(0, lambda: messagebox.showerror("Erro",
                "A biblioteca pdf2image não está disponível.\n"
                "Instale com: pip install pdf2image"))
            return

        if not PYTESSERACT_AVAILABLE:
            self.root.after(0, lambda: messagebox.showerror("Erro",
                "A biblioteca pytesseract não está disponível.\n"
                "Instale com: pip install pytesseract\n\n"
                "Instale também o Tesseract OCR:\n"
                "github.com/tesseract-ocr/tesseract"))
            return

        try:
            self.root.after(0, lambda: self.btn_ocr.config(state="disabled"))
            arquivo_entrada = self.ocr_arquivo_entrada.get()
            arquivo_saida = self.ocr_arquivo_saida.get()

            self._ocr_status("Convertendo PDF em imagens...", 0)

            poppler_path = self._get_poppler_path()
            kwargs = {'dpi': 300}
            if poppler_path:
                kwargs['poppler_path'] = poppler_path

            images = convert_from_path(arquivo_entrada, **kwargs)
            total_paginas = len(images)

            pdf_pages = []
            for i, img in enumerate(images):
                self._ocr_status(
                    f"Processando página {i + 1} de {total_paginas}...",
                    int((i / total_paginas) * 90))

                pdf_bytes = pytesseract.image_to_pdf_or_hocr(img, extension='pdf', lang='por')
                pdf_pages.append(io.BytesIO(pdf_bytes))

            self._ocr_status("Salvando PDF pesquisável...", 90)

            merger = PdfMerger()
            for page in pdf_pages:
                merger.append(page)
            merger.write(arquivo_saida)
            merger.close()

            self._ocr_status("PDF com OCR gerado com sucesso!", 100)
            self.root.after(0, lambda: messagebox.showinfo("Sucesso",
                f"PDF pesquisável gerado com sucesso!\nArquivo salvo em: {arquivo_saida}"))

        except Exception as e:
            msg = str(e)
            self._ocr_status("Erro durante o processo de OCR")
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro ao realizar OCR: {msg}"))

        finally:
            self.root.after(0, self._ocr_atualizar_botao)

    # ── Aba: Comprimir PDF ──────────────────────────────────────────────────

    def criar_aba_comprimir(self):
        aba = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(aba, text="🗜️ Comprimir PDF")

        aba.columnconfigure(1, weight=1)

        ttk.Label(aba, text="🗜️ Comprimir PDF", style='Title.TLabel').grid(
            row=0, column=0, columnspan=3, pady=(0, 20))

        ttk.Label(aba, text="📄 PDF de entrada:", style='Heading.TLabel').grid(
            row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(aba, textvariable=self.compress_arquivo_entrada, state="readonly", width=50).grid(
            row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(aba, text="Procurar", command=self.compress_selecionar_entrada).grid(
            row=1, column=2, pady=5)

        ttk.Label(aba, text="💾 Salvar como:", style='Heading.TLabel').grid(
            row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(aba, textvariable=self.compress_arquivo_saida, state="readonly", width=50).grid(
            row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(aba, text="Procurar", command=self.compress_selecionar_saida).grid(
            row=2, column=2, pady=5)

        nivel_frame = ttk.LabelFrame(aba, text="Nível de compressão", padding="10")
        nivel_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(15, 5))

        ttk.Radiobutton(nivel_frame, text="Baixa  (maior qualidade — DPI 200, qualidade 85%)",
                        variable=self.compress_nivel, value="baixa").grid(
            row=0, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(nivel_frame, text="Média  (equilibrado — DPI 150, qualidade 70%)",
                        variable=self.compress_nivel, value="media").grid(
            row=1, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(nivel_frame, text="Alta   (menor arquivo — DPI 100, qualidade 50%)",
                        variable=self.compress_nivel, value="alta").grid(
            row=2, column=0, sticky=tk.W, pady=2)

        self.compress_progress = ttk.Progressbar(aba, mode='determinate')
        self.compress_progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)

        self.compress_status_label = ttk.Label(aba, text="Selecione um PDF para comprimir")
        self.compress_status_label.grid(row=5, column=0, columnspan=3, pady=5)

        self.btn_comprimir = ttk.Button(aba, text="🗜️ Comprimir",
                                        command=self.comprimir_pdf_thread,
                                        state="disabled",
                                        style='Custom.TButton')
        self.btn_comprimir.grid(row=6, column=0, columnspan=3, pady=30, ipadx=20, ipady=10)

    def compress_selecionar_entrada(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o PDF para comprimir",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )
        if arquivo:
            self.compress_arquivo_entrada.set(arquivo)
            base, _ = os.path.splitext(arquivo)
            self.compress_arquivo_saida.set(base + "_comprimido.pdf")
            self._compress_atualizar_botao()

    def compress_selecionar_saida(self):
        arquivo = filedialog.asksaveasfilename(
            title="Salvar PDF comprimido como",
            defaultextension=".pdf",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )
        if arquivo:
            self.compress_arquivo_saida.set(arquivo)
            self._compress_atualizar_botao()

    def _compress_atualizar_botao(self):
        ok = bool(self.compress_arquivo_entrada.get() and self.compress_arquivo_saida.get())
        self.btn_comprimir.config(state="normal" if ok else "disabled")

    def comprimir_pdf_thread(self):
        thread = threading.Thread(target=self.comprimir_pdf)
        thread.daemon = True
        thread.start()

    def _status_ui(self, label, texto=None, progress=None, valor=None):
        """Atualiza um label de status e/ou progressbar na thread da UI (thread-safe).
        Use a partir de threads de trabalho — Tkinter não pode ser tocado fora da
        thread principal."""
        def upd():
            if texto is not None:
                label.config(text=texto)
            if progress is not None and valor is not None:
                progress.config(value=valor)
        self.root.after(0, upd)

    def _compress_status(self, texto=None, valor=None):
        """Atualiza status/progresso da compressão na thread da UI (thread-safe)."""
        self._status_ui(self.compress_status_label, texto, self.compress_progress, valor)

    def comprimir_pdf(self):
        if not PYMUPDF_AVAILABLE and not PDF2IMAGE_AVAILABLE:
            self.root.after(0, lambda: messagebox.showerror("Erro",
                "Nenhuma biblioteca de compressão disponível.\n"
                "Instale com: pip install PyMuPDF"))
            return

        # (dpi de rasterização para PDFs de imagem, qualidade JPEG)
        niveis = {
            "baixa": (200, 85),
            "media": (150, 70),
            "alta":  (100, 50),
        }
        dpi, qualidade = niveis[self.compress_nivel.get()]

        try:
            self.root.after(0, lambda: self.btn_comprimir.config(state="disabled"))
            arquivo_entrada = self.compress_arquivo_entrada.get()
            arquivo_saida = self.compress_arquivo_saida.get()

            self._compress_status("Analisando o PDF...", 0)

            # Decide a estratégia: PDF escaneado (imagem) vs texto/vetor
            eh_imagem = False
            if PYMUPDF_AVAILABLE:
                doc = fitz.open(arquivo_entrada)
                try:
                    eh_imagem = self._pdf_eh_imagem(doc)
                finally:
                    doc.close()

            if eh_imagem and PDF2IMAGE_AVAILABLE:
                self._comprimir_por_imagem(arquivo_entrada, arquivo_saida, dpi, qualidade)
            elif PYMUPDF_AVAILABLE:
                # PDF de texto/vetor: compressão sem perda (deflate + subset de fontes)
                max_dim = int(dpi * 8.27)  # limite ~ página A4 no DPI escolhido
                self._comprimir_texto(arquivo_entrada, arquivo_saida, qualidade, max_dim)
            else:
                self._comprimir_por_imagem(arquivo_entrada, arquivo_saida, dpi, qualidade)

            tamanho_original = os.path.getsize(arquivo_entrada) / 1024
            tamanho_final = os.path.getsize(arquivo_saida) / 1024

            if tamanho_final >= tamanho_original:
                shutil.copy2(arquivo_entrada, arquivo_saida)
                tamanho_final = tamanho_original
                self._compress_status(
                    f"Arquivo já otimizado — cópia salva sem alteração ({tamanho_original:.0f} KB)", 100)
                self.root.after(0, lambda: messagebox.showinfo("Sem redução",
                    f"O arquivo já está no menor tamanho possível com este método.\n"
                    f"O original foi copiado para: {arquivo_saida}"))
                return

            reducao = (1 - tamanho_final / tamanho_original) * 100

            self._compress_status(
                f"Concluído! {tamanho_original:.0f} KB → {tamanho_final:.0f} KB  ({reducao:.0f}% menor)", 100)
            self.root.after(0, lambda: messagebox.showinfo("Sucesso",
                f"PDF comprimido com sucesso!\n"
                f"Original: {tamanho_original:.0f} KB\n"
                f"Comprimido: {tamanho_final:.0f} KB  ({reducao:.0f}% menor)\n"
                f"Salvo em: {arquivo_saida}"))

        except Exception as e:
            msg = str(e)
            self._compress_status("Erro durante a compressão")
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro ao comprimir PDF: {msg}"))

        finally:
            self.root.after(0, self._compress_atualizar_botao)

    def _pdf_eh_imagem(self, doc):
        """Heurística: True se a maioria das páginas tem pouco/nenhum texto (PDF escaneado)."""
        paginas = len(doc)
        if paginas == 0:
            return False
        amostra = min(paginas, 10)
        com_texto = 0
        for i in range(amostra):
            try:
                if len(doc[i].get_text("text").strip()) > 50:
                    com_texto += 1
            except Exception:
                pass
        return com_texto < (amostra / 2)

    def _comprimir_texto(self, entrada, saida, qualidade, max_dim):
        """Compressão sem perda para PDFs de texto/vetor: recomprime imagens
        embutidas (best-effort), faz subset de fontes e regrava com deflate."""
        doc = fitz.open(entrada)
        try:
            self._compress_status("Otimizando imagens embutidas...", 30)
            self._recomprimir_imagens(doc, qualidade, max_dim)

            self._compress_status("Reduzindo fontes incorporadas...", 60)
            try:
                doc.subset_fonts()
            except Exception:
                pass  # subset não é suportado em todos os PDFs; segue sem ele

            self._compress_status("Salvando PDF comprimido...", 85)
            doc.save(saida, garbage=4, deflate=True, deflate_images=True,
                     deflate_fonts=True, clean=True)
        finally:
            doc.close()

    def _recomprimir_imagens(self, doc, qualidade, max_dim):
        """Recomprime/reduz imagens raster embutidas, preservando o texto vetorial.
        Best-effort: só substitui quando o resultado for menor, ignora imagens
        com transparência para não corromper o documento."""
        vistos = set()
        for page in doc:
            for img in page.get_images(full=True):
                xref = img[0]
                if xref in vistos:
                    continue
                vistos.add(xref)
                try:
                    base = doc.extract_image(xref)
                except Exception:
                    continue
                if base.get("smask", 0):
                    continue  # imagem com máscara/transparência — não mexe
                img_bytes = base.get("image", b"")
                if len(img_bytes) < 8 * 1024:
                    continue  # muito pequena, não compensa
                try:
                    pil = Image.open(io.BytesIO(img_bytes))
                    pil.load()
                    if pil.mode != "RGB":
                        pil = pil.convert("RGB")
                    if max(pil.size) > max_dim:
                        ratio = max_dim / max(pil.size)
                        pil = pil.resize(
                            (max(1, int(pil.width * ratio)), max(1, int(pil.height * ratio))),
                            Image.Resampling.LANCZOS)
                    buf = io.BytesIO()
                    pil.save(buf, format="JPEG", quality=qualidade, optimize=True)
                    novo = buf.getvalue()
                except Exception:
                    continue
                if len(novo) < len(img_bytes):
                    try:
                        page.replace_image(xref, stream=novo)
                    except Exception:
                        continue

    def _comprimir_por_imagem(self, arquivo_entrada, arquivo_saida, dpi, qualidade):
        """Rasteriza cada página em JPEG. Ideal para PDFs escaneados (imagem)."""
        poppler_path = self._get_poppler_path()
        kwargs = {'dpi': dpi}
        if poppler_path:
            kwargs['poppler_path'] = poppler_path

        self._compress_status("Convertendo páginas...")

        imagens_originais = convert_from_path(arquivo_entrada, **kwargs)
        total = len(imagens_originais)

        imagens_comprimidas = []
        for i, img in enumerate(imagens_originais):
            self._compress_status(
                f"Comprimindo página {i + 1} de {total}...", int((i / total) * 90))

            buf = io.BytesIO()
            img.convert("RGB").save(buf, format="JPEG", quality=qualidade, optimize=True)
            buf.seek(0)
            imagens_comprimidas.append(Image.open(buf).copy())

        self._compress_status("Salvando PDF comprimido...", 90)

        imagens_comprimidas[0].save(
            arquivo_saida,
            save_all=True,
            append_images=imagens_comprimidas[1:]
        )

    # ── Aba: PDF para Excel ─────────────────────────────────────────────────

    def criar_aba_pdf_excel(self):
        aba = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(aba, text="📊 PDF para Excel")

        aba.columnconfigure(1, weight=1)

        ttk.Label(aba, text="📊 PDF para Excel", style='Title.TLabel').grid(
            row=0, column=0, columnspan=3, pady=(0, 20))

        ttk.Label(aba, text="📄 PDF de entrada:", style='Heading.TLabel').grid(
            row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(aba, textvariable=self.excel_arquivo_entrada, state="readonly", width=50).grid(
            row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(aba, text="Procurar", command=self.excel_selecionar_entrada).grid(
            row=1, column=2, pady=5)

        ttk.Label(aba, text="💾 Salvar como:", style='Heading.TLabel').grid(
            row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(aba, textvariable=self.excel_arquivo_saida, state="readonly", width=50).grid(
            row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(aba, text="Procurar", command=self.excel_selecionar_saida).grid(
            row=2, column=2, pady=5)

        modo_frame = ttk.LabelFrame(aba, text="Modo de extração", padding="10")
        modo_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(15, 5))

        ttk.Radiobutton(
            modo_frame,
            text="Tabelas  (detecta e extrai tabelas estruturadas — ideal para PDFs com bordas)",
            variable=self.excel_modo, value="tabelas"
        ).grid(row=0, column=0, sticky=tk.W, pady=2)

        ttk.Radiobutton(
            modo_frame,
            text="Todo o texto  (extrai linha por linha — ideal para PDFs sem tabelas definidas)",
            variable=self.excel_modo, value="texto"
        ).grid(row=1, column=0, sticky=tk.W, pady=2)

        ttk.Label(aba,
                  text="ℹ️  Requer: pip install pdfplumber openpyxl",
                  foreground='#666666').grid(row=4, column=0, columnspan=3, pady=(10, 0))

        self.excel_progress = ttk.Progressbar(aba, mode='determinate')
        self.excel_progress.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)

        self.excel_status_label = ttk.Label(aba, text="Selecione um PDF para converter em Excel")
        self.excel_status_label.grid(row=6, column=0, columnspan=3, pady=5)

        self.btn_excel = ttk.Button(aba, text="📊 Converter para Excel",
                                    command=self.excel_converter_thread,
                                    state="disabled",
                                    style='Custom.TButton')
        self.btn_excel.grid(row=7, column=0, columnspan=3, pady=30, ipadx=20, ipady=10)

    def excel_selecionar_entrada(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o PDF para converter",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )
        if arquivo:
            self.excel_arquivo_entrada.set(arquivo)
            base, _ = os.path.splitext(arquivo)
            self.excel_arquivo_saida.set(base + ".xlsx")
            self._excel_atualizar_botao()

    def excel_selecionar_saida(self):
        arquivo = filedialog.asksaveasfilename(
            title="Salvar Excel como",
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")]
        )
        if arquivo:
            self.excel_arquivo_saida.set(arquivo)
            self._excel_atualizar_botao()

    def _excel_atualizar_botao(self):
        ok = bool(self.excel_arquivo_entrada.get() and self.excel_arquivo_saida.get())
        self.btn_excel.config(state="normal" if ok else "disabled")

    def excel_converter_thread(self):
        thread = threading.Thread(target=self.excel_converter)
        thread.daemon = True
        thread.start()

    def _extrair_por_coordenadas(self, pagina):
        """Extrai palavras agrupando por linha (Y) e separando colunas por gaps no eixo X.

        Detecta posições fixas de colunas a partir das linhas com data para dividir
        corretamente as linhas de continuação (ex: 'Negócios 988 BENEFICIÁRIO').
        """
        import re
        DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}')

        words = pagina.extract_words(x_tolerance=3, y_tolerance=3)
        if not words:
            return []

        grupos = {}
        for w in words:
            y_key = None
            for k in grupos:
                if abs(k - w['top']) <= 5:
                    y_key = k
                    break
            if y_key is None:
                y_key = w['top']
                grupos[y_key] = []
            grupos[y_key].append(w)

        # detecta onde as colunas começam nas linhas com data (gap > 10px entre grupos)
        col_boundaries = []
        for y in sorted(grupos.keys()):
            ws_linha = sorted(grupos[y], key=lambda w: w['x0'])
            if not DATE_RE.match(ws_linha[0]['text']):
                continue
            boundaries = []
            prev_x1 = ws_linha[0]['x1']
            for w in ws_linha[1:]:
                if w['x0'] - prev_x1 > 10:
                    boundaries.append(w['x0'])
                prev_x1 = w['x1']
            if boundaries:
                col_boundaries.append(boundaries)

        # média das posições de início de cada coluna
        avg_boundaries = []
        if col_boundaries:
            max_cols = max(len(b) for b in col_boundaries)
            for i in range(max_cols):
                vals = [b[i] for b in col_boundaries if i < len(b)]
                avg_boundaries.append(sum(vals) / len(vals))

        resultado = []
        for y in sorted(grupos.keys()):
            ws_linha = sorted(grupos[y], key=lambda w: w['x0'])
            colunas = []
            texto_col = ws_linha[0]['text']
            x_fim = ws_linha[0]['x1']

            for w in ws_linha[1:]:
                gap = w['x0'] - x_fim
                near_boundary = any(abs(w['x0'] - b) < 5 for b in avg_boundaries)
                if gap > 15 or (near_boundary and gap > 3):
                    colunas.append(texto_col)
                    texto_col = w['text']
                else:
                    texto_col += ' ' + w['text']
                x_fim = w['x1']

            colunas.append(texto_col)
            resultado.append(colunas)

        return resultado

    def _mesclar_continuacoes(self, linhas):
        """Mescla linhas de continuação na linha principal anterior.

        Se o documento tiver datas (DD/MM/AAAA), usa esse padrão para identificar
        linhas principais. Caso contrário, usa o máximo de colunas como referência.
        """
        import re
        if not linhas:
            return linhas

        DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}')
        tem_datas = any(DATE_RE.match(str(l[0])) for l in linhas if l and l[0])

        resultado = []
        dentro_transacoes = False
        n_colunas_principal = 0

        if not tem_datas:
            # sem datas: usa o máximo de colunas como limiar
            esperado = max((len(l) for l in linhas if l), default=1)
            limiar = max(2, round(esperado * 0.6))

        for linha in linhas:
            if not linha:
                continue

            col0 = str(linha[0]) if linha[0] else ''

            if tem_datas:
                if DATE_RE.match(col0):
                    dentro_transacoes = True
                    # separa a data do restante (Local) que vem colado na col 0
                    data = col0[:10]
                    local = col0[10:].strip()
                    resto = [str(c) if c is not None else '' for c in linha[1:]]
                    row = [data, local] + resto
                    resultado.append(row)
                    n_colunas_principal = max(n_colunas_principal, len(row))
                elif dentro_transacoes and resultado:
                    # linha de continuação: agrega ao Local (col 1) e complemento como col extra
                    cont_local = col0
                    cont_extra = str(linha[1]) if len(linha) > 1 and linha[1] else ''
                    if cont_local:
                        resultado[-1][1] = (resultado[-1][1] + ' ' + cont_local).strip()
                    if cont_extra:
                        while len(resultado[-1]) < n_colunas_principal + 1:
                            resultado[-1].append('')
                        idx = n_colunas_principal
                        existing = resultado[-1][idx]
                        resultado[-1][idx] = existing + (' | ' if existing else '') + cont_extra
                else:
                    # metadados pré-transação: inclui como está
                    resultado.append([str(c) if c is not None else '' for c in linha])
            else:
                if len(linha) >= limiar or not resultado:
                    resultado.append([str(c) if c is not None else '' for c in linha])
                else:
                    cont = ' '.join(str(c) for c in linha if c)
                    if cont:
                        resultado[-1][0] = resultado[-1][0] + ' ' + cont

        return resultado

    def _extrair_pagina_tabela(self, pagina):
        """Tenta tabela estruturada; se dados ficarem numa só coluna, usa coordenadas."""
        tabelas = pagina.extract_tables()
        if tabelas:
            linhas = [linha for tabela in tabelas for linha in tabela]
            n_colunas = max((len(r) for r in linhas), default=0)
            total = sum(len(r) for r in linhas)
            nulos = sum(1 for r in linhas for c in r if c is None)
            if n_colunas > 1 and total > 0 and nulos / total < 0.5:
                return self._mesclar_continuacoes(linhas)

        return self._mesclar_continuacoes(self._extrair_por_coordenadas(pagina))

    def _excel_status(self, texto=None, valor=None):
        self._status_ui(self.excel_status_label, texto, self.excel_progress, valor)

    def excel_converter(self):
        if not PDF_EXCEL_AVAILABLE:
            self.root.after(0, lambda: messagebox.showerror("Erro",
                "As bibliotecas necessárias não estão instaladas.\n"
                "Execute: pip install pdfplumber openpyxl"))
            return

        try:
            self.root.after(0, lambda: self.btn_excel.config(state="disabled"))
            arquivo_entrada = self.excel_arquivo_entrada.get()
            arquivo_saida = self.excel_arquivo_saida.get()
            modo = self.excel_modo.get()

            self._excel_status("Abrindo PDF...", 0)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dados"

            cabecalho_font = Font(bold=True, color="FFFFFF")
            cabecalho_fill = PatternFill(fill_type="solid", fgColor="2C3E50")
            centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

            linha_atual = 1
            linhas_escritas = 0
            linha_cabecalho = None

            with pdfplumber.open(arquivo_entrada) as pdf:
                total = len(pdf.pages)

                for i, pagina in enumerate(pdf.pages):
                    self._excel_status(
                        f"Processando página {i + 1} de {total}...", int((i / total) * 95))

                    if modo == "tabelas":
                        linhas = self._extrair_pagina_tabela(pagina)
                        for linha in linhas:
                            celulas_com_valor = [c for c in linha if c and str(c).strip()]
                            if not celulas_com_valor:
                                continue
                            # pula repetições do cabeçalho nas páginas seguintes
                            if linha_cabecalho is not None and linha == linha_cabecalho:
                                continue
                            eh_cabecalho = linha_cabecalho is None
                            if eh_cabecalho:
                                linha_cabecalho = linha
                            for c_idx, celula in enumerate(linha):
                                valor = str(celula).strip() if celula is not None else ""
                                cell = ws.cell(row=linha_atual, column=c_idx + 1, value=valor)
                                if eh_cabecalho:
                                    cell.font = cabecalho_font
                                    cell.fill = cabecalho_fill
                                    cell.alignment = centro
                            linha_atual += 1
                            linhas_escritas += 1

                    else:
                        texto = pagina.extract_text()
                        if texto and texto.strip():
                            for linha_txt in texto.splitlines():
                                if linha_txt.strip():
                                    ws.cell(row=linha_atual, column=1, value=linha_txt)
                                    linha_atual += 1
                                    linhas_escritas += 1

            if linhas_escritas == 0:
                self._excel_status("Nenhum dado encontrado")
                self.root.after(0, lambda: messagebox.showwarning("Sem dados",
                    "Nenhum dado foi encontrado no PDF com o modo selecionado.\n"
                    "Tente o modo 'Todo o texto' se o PDF não tiver tabelas detectáveis."))
                return

            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) if c.value else 0 for c in col), default=0
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            self._excel_status("Salvando arquivo Excel...", 97)

            wb.save(arquivo_saida)

            self._excel_status(f"Concluído! {linhas_escritas} linha(s) extraída(s).", 100)
            self.root.after(0, lambda: messagebox.showinfo("Sucesso",
                f"Excel gerado com sucesso!\n"
                f"Linhas extraídas: {linhas_escritas}\n"
                f"Salvo em: {arquivo_saida}"))

        except Exception as e:
            msg = str(e)
            self._excel_status("Erro durante a conversão")
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro ao converter para Excel: {msg}"))

        finally:
            self.root.after(0, self._excel_atualizar_botao)

    # ── Helpers ─────────────────────────────────────────────────────────────

    def _get_poppler_path(self):
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
        poppler_path = os.path.join(base_path, 'poppler_bin')
        return poppler_path if os.path.exists(poppler_path) else None

    # ── Juntar PDFs: lógica existente ───────────────────────────────────────

    def selecionar_pasta(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta com os PDFs")
        if pasta:
            self.pasta_origem.set(pasta)
            self.listar_pdfs_da_pasta()

    def selecionar_destino(self):
        arquivo = filedialog.asksaveasfilename(
            title="Salvar PDF unificado como",
            defaultextension=".pdf",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )
        if arquivo:
            self.arquivo_destino.set(arquivo)
            self.verificar_botao_juntar()

    def adicionar_arquivos(self):
        arquivos = filedialog.askopenfilenames(
            title="Selecione arquivos PDF",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )

        for arquivo in arquivos:
            nome_arquivo = os.path.basename(arquivo)
            caminho_completo = arquivo

            if caminho_completo not in [item['caminho'] for item in self.arquivos_pdf]:
                self.arquivos_pdf.append({
                    'nome': nome_arquivo,
                    'caminho': caminho_completo
                })

        self.atualizar_listbox()
        self.verificar_botao_juntar()

    def listar_pdfs_da_pasta(self):
        pasta = self.pasta_origem.get()
        if not pasta:
            return

        try:
            self.arquivos_pdf.clear()

            arquivos_encontrados = []
            for arquivo in os.listdir(pasta):
                if arquivo.lower().endswith('.pdf'):
                    arquivos_encontrados.append(arquivo)

            arquivos_encontrados.sort()

            for arquivo in arquivos_encontrados:
                self.arquivos_pdf.append({
                    'nome': arquivo,
                    'caminho': os.path.join(pasta, arquivo)
                })

            self.atualizar_listbox()

            if self.arquivos_pdf:
                self.status_label.config(text=f"{len(self.arquivos_pdf)} arquivo(s) PDF encontrado(s)")
            else:
                self.status_label.config(text="Nenhum arquivo PDF encontrado na pasta")

            self.verificar_botao_juntar()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao listar arquivos: {str(e)}")
            self.status_label.config(text="Erro ao acessar a pasta")

    def atualizar_listbox(self):
        self.listbox.delete(0, tk.END)
        for i, arquivo in enumerate(self.arquivos_pdf):
            self.listbox.insert(tk.END, f"{i + 1:02d}. {arquivo['nome']}")

    def remover_selecionados(self):
        selecionados = self.listbox.curselection()
        if not selecionados:
            messagebox.showwarning("Aviso", "Selecione pelo menos um arquivo para remover")
            return

        for i in reversed(selecionados):
            del self.arquivos_pdf[i]

        self.atualizar_listbox()
        self.limpar_preview()
        self.verificar_botao_juntar()
        self.atualizar_botoes_controle()

    def mover_para_cima(self):
        selecionados = self.listbox.curselection()
        if len(selecionados) != 1:
            messagebox.showwarning("Aviso", "Selecione apenas um arquivo para mover")
            return

        indice = selecionados[0]
        if indice > 0:
            self.arquivos_pdf[indice], self.arquivos_pdf[indice - 1] = \
                self.arquivos_pdf[indice - 1], self.arquivos_pdf[indice]
            self.atualizar_listbox()
            self.listbox.selection_set(indice - 1)

    def mover_para_baixo(self):
        selecionados = self.listbox.curselection()
        if len(selecionados) != 1:
            messagebox.showwarning("Aviso", "Selecione apenas um arquivo para mover")
            return

        indice = selecionados[0]
        if indice < len(self.arquivos_pdf) - 1:
            self.arquivos_pdf[indice], self.arquivos_pdf[indice + 1] = \
                self.arquivos_pdf[indice + 1], self.arquivos_pdf[indice]
            self.atualizar_listbox()
            self.listbox.selection_set(indice + 1)

    def on_listbox_select(self, event):
        self.atualizar_botoes_controle()
        self.mostrar_preview()

    def atualizar_botoes_controle(self):
        selecionados = self.listbox.curselection()

        if selecionados:
            self.btn_remover.config(state="normal")
        else:
            self.btn_remover.config(state="disabled")

        if len(selecionados) == 1:
            indice = selecionados[0]
            self.btn_subir.config(state="normal" if indice > 0 else "disabled")
            self.btn_descer.config(state="normal" if indice < len(self.arquivos_pdf) - 1 else "disabled")
        else:
            self.btn_subir.config(state="disabled")
            self.btn_descer.config(state="disabled")

    def mostrar_preview(self):
        if not PDF2IMAGE_AVAILABLE:
            self.preview_info.config(text="Pré-visualização não disponível (pdf2image não instalado)")
            return

        selecionados = self.listbox.curselection()
        if len(selecionados) != 1:
            self.limpar_preview()
            return

        indice = selecionados[0]
        arquivo = self.arquivos_pdf[indice]

        # Cada chamada recebe um token; resultados antigos (cliques rápidos) são descartados
        self.preview_token += 1
        token = self.preview_token

        # Se já temos a imagem em cache, exibe imediatamente
        cache_key = arquivo['caminho']
        if cache_key in self.preview_cache:
            self._exibir_preview(self.preview_cache[cache_key], arquivo['nome'])
            return

        self.preview_info.config(text=f"Carregando pré-visualização de: {arquivo['nome']}")

        thread = threading.Thread(
            target=self._render_preview_thread,
            args=(arquivo['caminho'], arquivo['nome'], token),
            daemon=True
        )
        thread.start()

    def _render_preview_thread(self, caminho, nome, token):
        try:
            poppler_path = self._get_poppler_path()
            kwargs = {'first_page': 1, 'last_page': 1, 'dpi': 100}
            if poppler_path:
                kwargs['poppler_path'] = poppler_path

            images = convert_from_path(caminho, **kwargs)
            img = images[0] if images else None

            # Volta para a thread da UI para atualizar o canvas
            self.root.after(0, lambda: self._preview_concluido(img, caminho, nome, token))
        except Exception as e:
            self.root.after(0, lambda: self._preview_erro(str(e), token))

    def _preview_concluido(self, img, caminho, nome, token):
        # Ignora resultados de seleções já substituídas
        if token != self.preview_token:
            return

        if img is None:
            self.preview_info.config(text="Erro ao carregar pré-visualização")
            return

        self.preview_cache[caminho] = img
        self._exibir_preview(img, nome)

    def _preview_erro(self, msg, token):
        if token != self.preview_token:
            return
        self.preview_info.config(text=f"Erro na pré-visualização: {msg}")
        self.limpar_preview()

    def _exibir_preview(self, img, nome):
        canvas_width = self.preview_canvas.winfo_width()
        canvas_height = self.preview_canvas.winfo_height()

        if canvas_width > 1 and canvas_height > 1:
            img_ratio = img.width / img.height
            canvas_ratio = canvas_width / canvas_height

            if img_ratio > canvas_ratio:
                new_width = min(canvas_width - 20, img.width)
                new_height = int(new_width / img_ratio)
            else:
                new_height = min(canvas_height - 20, img.height)
                new_width = int(new_height * img_ratio)

            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)

        self.preview_image = ImageTk.PhotoImage(img)

        self.preview_canvas.delete("all")
        self.preview_canvas.create_image(
            self.preview_canvas.winfo_width() // 2,
            self.preview_canvas.winfo_height() // 2,
            image=self.preview_image,
            anchor=tk.CENTER
        )
        self.preview_canvas.configure(scrollregion=self.preview_canvas.bbox("all"))
        self.preview_info.config(text=f"Pré-visualização: {nome}")

    def limpar_preview(self):
        self.preview_token += 1  # descarta qualquer render em andamento
        self.preview_canvas.delete("all")
        self.preview_image = None
        self.preview_info.config(text="Selecione um arquivo para ver a pré-visualização")

    def verificar_botao_juntar(self):
        if self.arquivo_destino.get() and len(self.arquivos_pdf) > 0:
            self.btn_juntar.config(state="normal")
        else:
            self.btn_juntar.config(state="disabled")

    def juntar_pdfs_thread(self):
        thread = threading.Thread(target=self.juntar_pdfs)
        thread.daemon = True
        thread.start()

    def _juntar_status(self, texto=None, valor=None):
        self._status_ui(self.status_label, texto, self.progress, valor)

    def juntar_pdfs(self):
        try:
            self.root.after(0, lambda: self.btn_juntar.config(state="disabled"))
            self._juntar_status("Juntando PDFs...", 0)

            arquivo_destino = self.arquivo_destino.get()
            total_arquivos = len(self.arquivos_pdf)

            merger = PdfMerger()

            for i, arquivo_info in enumerate(self.arquivos_pdf):
                try:
                    merger.append(arquivo_info['caminho'])
                    progresso = int((i + 1) / total_arquivos * 100)
                    self._juntar_status(f"Processando: {arquivo_info['nome']}", progresso)
                except Exception as e:
                    nome, msg = arquivo_info['nome'], str(e)
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Aviso", f"Erro ao processar {nome}: {msg}"))
                    continue

            self._juntar_status("Salvando arquivo final...")
            merger.write(arquivo_destino)
            merger.close()

            self._juntar_status("PDFs unidos com sucesso!", 100)
            self.root.after(0, lambda: messagebox.showinfo("Sucesso",
                f"PDFs unidos com sucesso!\nArquivo salvo em: {arquivo_destino}"))

        except Exception as e:
            msg = str(e)
            self._juntar_status("Erro durante o processo")
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro ao juntar PDFs: {msg}"))

        finally:
            self.root.after(0, lambda: self.btn_juntar.config(state="normal"))


def main():
    root = tk.Tk()
    app = PDFMergerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
